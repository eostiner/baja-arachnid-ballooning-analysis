#!/usr/bin/env python3
"""
Publication QC for Baja arachnid ballooning analyses.

This script never invents data or silently substitutes a legacy binary trait table.
It reads the retained 25-km genus-by-cell incidence matrix, cell-to-band lookup,
and reviewed D1-D4/N0 trait table, then performs paired equal-cell resampling.

Primary outputs:
  * input provenance and SHA-256 hashes
  * exact trait audit (D1-D4/N0; D4 excluded)
  * equal-cell richness by latitude band and trait
  * pairwise total Jaccard dissimilarity
  * Baselga Jaccard replacement and nestedness-resultant partition
  * Simpson replacement and Sorensen nestedness-resultant partition
  * paired ballooning-minus-non-ballooning contrasts
  * iNEXT incidence-frequency input derived from the same frozen matrix
  * publication-QC figures and caption draft using calculated values only
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import random
import re
import statistics
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from itertools import combinations
from pathlib import Path
from typing import Any, Iterable, Sequence

SCRIPT_VERSION = "2026-07-23-publication-qc-v1.2"
DEFAULT_ITERATIONS = 5000
DEFAULT_SEED = 20260723
DEFAULT_EXPECTED_EQUAL_CELLS = 22
BAND_ORDER = ["23-24N", "24-26N", "26-28N", "28-30N", "30-32N"]
BAND_LABEL = {
    "23-24N": "23–24°N",
    "24-26N": "24–26°N",
    "26-28N": "26–28°N",
    "28-30N": "28–30°N",
    "30-32N": "30–32°N",
}
ADJACENT_PAIRS = list(zip(BAND_ORDER[:-1], BAND_ORDER[1:]))
TRAIT_ORDER = ["ballooning", "non_ballooning"]
TRAIT_LABEL = {
    "ballooning": "Ballooning-capable",
    "non_ballooning": "Non-ballooning",
}
METRICS = [
    "jaccard_total",
    "jaccard_turnover",
    "jaccard_nestedness",
    "sorensen_total",
    "simpson_replacement",
    "sorensen_nestedness",
]


@dataclass(frozen=True)
class Inputs:
    project_root: Path
    incidence_matrix: Path
    cell_lookup: Path
    trait_table: Path
    occurrence_table: Path | None


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for block in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def write_csv(path: Path, rows: Iterable[dict[str, Any]], fields: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(fields), extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({f: row.get(f, "") for f in fields})
    tmp.replace(path)


def read_delimited(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    delimiter = "\t" if path.suffix.lower() in {".tsv", ".tab"} else ","
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh, delimiter=delimiter)
        if not reader.fieldnames:
            raise ValueError(f"No header found in {path}")
        return list(reader.fieldnames), [dict(row) for row in reader]


def read_table(path: Path) -> tuple[list[str], list[dict[str, str]], str | None]:
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
        fields, rows = read_delimited(path)
        return fields, rows, None
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read the authoritative Excel trait workbook.") from exc
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    preferred = "Genus_Trait_Master_267"
    sheet = preferred if preferred in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    iterator = ws.iter_rows(values_only=True)
    try:
        header_raw = next(iterator)
    except StopIteration as exc:
        raise ValueError(f"Empty workbook sheet {sheet!r} in {path}") from exc
    fields = [str(v).strip() if v is not None else "" for v in header_raw]
    if not any(fields):
        raise ValueError(f"Blank header in {path} sheet {sheet}")
    rows: list[dict[str, str]] = []
    for values in iterator:
        row = {field: "" if value is None else str(value).strip() for field, value in zip(fields, values) if field}
        if any(str(v).strip() for v in row.values()):
            rows.append(row)
    return [f for f in fields if f], rows, sheet


def find_field(fields: Sequence[str], candidates: Sequence[str], label: str, required: bool = True) -> str | None:
    lookup = {f.casefold(): f for f in fields}
    for candidate in candidates:
        if candidate.casefold() in lookup:
            return lookup[candidate.casefold()]
    if required:
        raise ValueError(f"Could not identify {label}. Tried {candidates}. Available: {fields}")
    return None


def first_existing(paths: Sequence[Path], label: str, required: bool = True) -> Path | None:
    for path in paths:
        path = path.expanduser()
        if path.is_file():
            return path.resolve()
    if required:
        raise FileNotFoundError(f"Could not find {label}. Tried:\n" + "\n".join(str(p.expanduser()) for p in paths))
    return None


def auto_project_root(explicit: str | None) -> Path:
    if explicit:
        root = Path(explicit).expanduser().resolve()
        if not root.is_dir():
            raise FileNotFoundError(f"Project root not found: {root}")
        return root
    candidates = [
        Path("~/Desktop/OLD BALLOONING/Baja_Ballooning_Pipeline"),
        Path("~/Desktop/BALLOONING Overflo/Baja_Ballooning_Pipeline"),
        Path("~/Desktop/Baja_Ballooning_Pipeline"),
        Path("~/Downloads/Baja_Ballooning_Pipeline"),
    ]
    for candidate in candidates:
        expanded = candidate.expanduser()
        if expanded.is_dir():
            return expanded.resolve()
    raise FileNotFoundError(
        "Could not auto-detect the Baja_Ballooning_Pipeline project root. "
        "Pass it as the first argument to RUN_PUBLICATION_QC.command."
    )


def discover_inputs(root: Path, args: argparse.Namespace) -> Inputs:
    analysis_ready = root / "ANALYSIS_READY_INPUTS"
    clean = root / "02_data_clean"
    incidence = Path(args.incidence).expanduser().resolve() if args.incidence else first_existing([
        analysis_ready / "02_incidence_matrices_25km/10_biodiversity_final_genus_by_grid25km_incidence.csv",
        clean / "08_grid25km_incidence/10_biodiversity_final_genus_by_grid25km_incidence.csv",
    ], "primary 25-km genus incidence matrix")
    lookup = Path(args.cell_lookup).expanduser().resolve() if args.cell_lookup else first_existing([
        analysis_ready / "04_spatial_reference/10_common_grid25km_cell_lookup.csv",
        clean / "08_grid25km_incidence/10_common_grid25km_cell_lookup.csv",
    ], "25-km cell lookup")
    trait = Path(args.traits).expanduser().resolve() if args.traits else first_existing([
        # Prefer sources that explicitly contain D1/D2/D3/D4/N0 for every matrix genus.
        analysis_ready / "03_trait_tables/07_reviewed_genus_trait_lookup_final_REAUDITED_COMPLETE.csv",
        analysis_ready / "03_trait_tables/07_reviewed_genus_trait_lookup_normalized.csv",
        root / "USE_GOOD_BalloonID_Baja_Arachnid_GenusSpecies_Long_D1_D4_AUTHORITATIVE.xlsx",
        analysis_ready / "03_trait_tables/07_reviewed_genus_trait_lookup_final.csv",
        clean / "07_final_trait_merge/07_reviewed_genus_trait_lookup_final_REAUDITED_COMPLETE.csv",
        clean / "07_final_trait_merge/07_reviewed_genus_trait_lookup_final.csv",
    ], "reviewed D1-D4/N0 trait table")
    occurrence = first_existing([
        clean / "05_final_qc_flags/05_biodiversity_final_records.tsv",
        clean / "05_final_qc_flags/05_biodiversity_final_records.csv",
    ], "final-QC occurrence table", required=False)
    assert incidence is not None and lookup is not None and trait is not None
    return Inputs(root, incidence, lookup, trait, occurrence)


def normalize_band(value: str) -> str | None:
    text = str(value or "").strip().upper()
    text = text.replace("–", "-").replace("—", "-").replace("°", "")
    text = re.sub(r"\s+", "", text)
    text = text.replace("NORTH", "N")
    m = re.search(r"(23\s*-\s*24|24\s*-\s*26|26\s*-\s*28|28\s*-\s*30|30\s*-\s*32)N?", text)
    return (m.group(1).replace(" ", "") + "N") if m else None


def read_incidence(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        try:
            header = next(reader)
        except StopIteration as exc:
            raise ValueError(f"Empty incidence matrix: {path}") from exc
        if len(header) < 2:
            raise ValueError(f"Incidence matrix has fewer than two columns: {path}")
        genus_field = header[0].strip()
        cells = [x.strip() for x in header[1:]]
        if any(not c for c in cells) or len(cells) != len(set(cells)):
            raise ValueError("Blank or duplicate grid-cell columns in incidence matrix.")
        genera: list[str] = []
        rows_by_genus: dict[str, list[int]] = {}
        for row_num, row in enumerate(reader, start=2):
            if len(row) != len(header):
                raise ValueError(f"Row {row_num} has {len(row)} columns; expected {len(header)}")
            genus = row[0].strip()
            if not genus:
                raise ValueError(f"Blank genus at row {row_num}")
            key = genus.casefold()
            if key in rows_by_genus:
                raise ValueError(f"Duplicate genus in incidence matrix: {genus}")
            try:
                values = [int(v) for v in row[1:]]
            except ValueError as exc:
                raise ValueError(f"Non-integer incidence value at row {row_num}") from exc
            if any(v not in (0, 1) for v in values):
                raise ValueError(f"Values other than 0/1 at row {row_num}")
            if not any(values):
                raise ValueError(f"Genus has zero occupancy in retained matrix: {genus}")
            genera.append(genus)
            rows_by_genus[key] = values
    cell_masks = [0] * len(cells)
    for gi, genus in enumerate(genera):
        bit = 1 << gi
        values = rows_by_genus[genus.casefold()]
        for ci, value in enumerate(values):
            if value:
                cell_masks[ci] |= bit
    return {"genus_field": genus_field, "genera": genera, "cells": cells, "cell_masks": cell_masks, "rows": rows_by_genus}


def read_cell_bands(path: Path, required_cells: Sequence[str]) -> tuple[dict[str, str], list[dict[str, Any]]]:
    fields, rows = read_delimited(path)
    cell_field = find_field(fields, ["grid_cell_id", "cell_id", "grid25km_cell_id"], "grid-cell ID")
    band_field = find_field(fields, ["centroid_latitude_band", "latitude_band", "lat_band"], "latitude band")
    mapping: dict[str, str] = {}
    audit: list[dict[str, Any]] = []
    for row in rows:
        cell = row.get(cell_field or "", "").strip()
        raw = row.get(band_field or "", "").strip()
        if not cell:
            continue
        band = normalize_band(raw)
        if band is None:
            audit.append({"cell_id": cell, "raw_band": raw, "normalized_band": "UNRESOLVED"})
            continue
        if cell in mapping and mapping[cell] != band:
            raise ValueError(f"Conflicting band assignments for cell {cell}")
        mapping[cell] = band
        audit.append({"cell_id": cell, "raw_band": raw, "normalized_band": band})
    missing = [c for c in required_cells if c not in mapping]
    if missing:
        raise ValueError(f"Cell lookup lacks {len(missing)} incidence cells. Examples: {missing[:10]}")
    invalid = sorted({mapping[c] for c in required_cells if mapping[c] not in BAND_ORDER})
    if invalid:
        raise ValueError(f"Cells assigned outside the five retained bands: {invalid}")
    return {c: mapping[c] for c in required_cells}, audit


EVIDENCE_RE = re.compile(r"(?<![A-Z0-9])(D[1-4]|N0|C3)(?![A-Z0-9])", re.I)


def parse_evidence(value: Any) -> str | None:
    text = str(value or "").strip().upper()
    if not text:
        return None
    tokens = {t.upper() for t in EVIDENCE_RE.findall(text)}
    if len(tokens) == 1:
        return next(iter(tokens))
    normalized = re.sub(r"[^a-z0-9]+", "", text.casefold())
    if normalized in {"ballooning", "ballooningc3", "primaryc3", "d1d2d3", "d1tod3"}:
        return "C3"
    if normalized in {"nonballooning", "nonballooningn0", "fixednonballooning", "reference"}:
        return "N0"
    if normalized in {"excluded", "excludedd4", "d4excluded"}:
        return "D4"
    return None


def load_traits(path: Path, genera: Sequence[str]) -> dict[str, Any]:
    fields, rows, sheet = read_table(path)
    genus_field = find_field(fields, ["genus", "analysis_genus", "trait_review_genus"], "trait genus")
    evidence_fields = [f for f in fields if f.casefold() in {
        "evidence_class", "final_evidence_class", "final_evidence_category", "exclusive_tier",
        "final_tier_for_current_build", "trait_class", "analysis_class", "primary_c3_group",
        "final_designation", "designation", "ballooning_evidence_tier"
    }]
    if not evidence_fields:
        raise ValueError(
            "Trait table lacks explicit D1/D2/D3/D4/N0 or C3/N0 fields. "
            "A legacy yes/no ballooning column is intentionally rejected."
        )
    lookup: dict[str, dict[str, Any]] = {}
    duplicate_rows: list[str] = []
    for row in rows:
        genus = row.get(genus_field or "", "").strip()
        if not genus:
            continue
        key = genus.casefold()
        parsed_by_field = {f: parse_evidence(row.get(f, "")) for f in evidence_fields}
        parsed = {v for v in parsed_by_field.values() if v is not None}
        # Treat C3 and D1-D3 as mutually compatible primary ballooning designations.
        collapsed = {"C3" if v in {"D1", "D2", "D3", "C3"} else v for v in parsed}
        if len(collapsed) > 1:
            raise ValueError(f"Contradictory trait coding for {genus}: {parsed_by_field}")
        evidence = None
        for f in evidence_fields:
            v = parsed_by_field[f]
            if v in {"D1", "D2", "D3", "D4", "N0"}:
                evidence = v
                break
        if evidence is None and parsed:
            evidence = next(iter(parsed))
        if key in lookup:
            duplicate_rows.append(genus)
        else:
            lookup[key] = {"genus": genus, "evidence": evidence, "field_values": parsed_by_field}
    if duplicate_rows:
        raise ValueError(f"Duplicate genera in trait table: {duplicate_rows[:20]}")
    normalized_rows: list[dict[str, Any]] = []
    missing, unresolved = [], []
    masks = {"ballooning": 0, "non_ballooning": 0, "excluded_D4": 0}
    evidence_counts: Counter[str] = Counter()
    for gi, genus in enumerate(genera):
        item = lookup.get(genus.casefold())
        if item is None:
            missing.append(genus)
            continue
        ev = item["evidence"]
        if ev is None:
            unresolved.append(genus)
            continue
        if ev in {"D1", "D2", "D3", "C3"}:
            trait = "ballooning"
        elif ev == "N0":
            trait = "non_ballooning"
        elif ev == "D4":
            trait = "excluded_D4"
        else:
            unresolved.append(genus)
            continue
        masks[trait] |= 1 << gi
        evidence_counts[ev] += 1
        normalized_rows.append({"genus": genus, "evidence_class": ev, "analysis_class": trait})
    if missing:
        raise ValueError(f"Genera missing from trait table ({len(missing)}): {missing[:20]}")
    if unresolved:
        raise ValueError(
            f"Genera with unresolved explicit trait coding ({len(unresolved)}): {unresolved[:20]}. "
            f"Explicit fields detected: {evidence_fields}. This usually means a partial legacy "
            "07_reviewed_genus_trait_lookup_final.csv was selected. Use "
            "07_reviewed_genus_trait_lookup_final_REAUDITED_COMPLETE.csv or the authoritative "
            "D1-D4 workbook containing final_tier_for_current_build."
        )
    full_mask = (1 << len(genera)) - 1
    if (masks["ballooning"] | masks["non_ballooning"] | masks["excluded_D4"]) != full_mask:
        raise RuntimeError("Trait masks do not cover the full incidence-matrix genus set.")
    if masks["ballooning"] & masks["non_ballooning"]:
        raise RuntimeError("Ballooning and non-ballooning masks overlap.")
    return {
        "sheet": sheet,
        "fields_used": evidence_fields,
        "masks": masks,
        "evidence_counts": dict(sorted(evidence_counts.items())),
        "normalized_rows": normalized_rows,
    }


def union_masks(cell_masks: Sequence[int], indices: Sequence[int]) -> int:
    value = 0
    for idx in indices:
        value |= cell_masks[idx]
    return value


def baselga_pair(mask_a: int, mask_b: int) -> dict[str, float | int]:
    """Pairwise Baselga-family partitions from a, b, c incidence counts.

    a: shared genera
    b: unique to assemblage 1
    c: unique to assemblage 2

    Jaccard family:
      beta_jac = (b+c)/(a+b+c)
      beta_jtu = 2*min(b,c)/(a+2*min(b,c))
      beta_jne = beta_jac-beta_jtu

    Sorensen family:
      beta_sor = (b+c)/(2a+b+c)
      beta_sim = min(b,c)/(a+min(b,c))
      beta_sne = beta_sor-beta_sim
    """
    a = (mask_a & mask_b).bit_count()
    b = (mask_a & ~mask_b).bit_count()
    c = (mask_b & ~mask_a).bit_count()
    m = min(b, c)
    d_jac = a + b + c
    d_jtu = a + 2 * m
    d_sor = 2 * a + b + c
    d_sim = a + m
    beta_jac = (b + c) / d_jac if d_jac else math.nan
    beta_jtu = (2 * m) / d_jtu if d_jtu else math.nan
    beta_jne = beta_jac - beta_jtu if not math.isnan(beta_jac) and not math.isnan(beta_jtu) else math.nan
    beta_sor = (b + c) / d_sor if d_sor else math.nan
    beta_sim = m / d_sim if d_sim else math.nan
    beta_sne = beta_sor - beta_sim if not math.isnan(beta_sor) and not math.isnan(beta_sim) else math.nan
    return {
        "shared_genera": a,
        "unique_to_band_1": b,
        "unique_to_band_2": c,
        "jaccard_total": beta_jac,
        "jaccard_turnover": max(0.0, beta_jtu) if not math.isnan(beta_jtu) else math.nan,
        "jaccard_nestedness": max(0.0, beta_jne) if not math.isnan(beta_jne) else math.nan,
        "sorensen_total": beta_sor,
        "simpson_replacement": beta_sim,
        "sorensen_nestedness": max(0.0, beta_sne) if not math.isnan(beta_sne) else math.nan,
    }


def percentile(values: Sequence[float], p: float) -> float:
    vals = sorted(v for v in values if not math.isnan(v))
    if not vals:
        return math.nan
    if len(vals) == 1:
        return vals[0]
    pos = (len(vals) - 1) * p
    lo, hi = math.floor(pos), math.ceil(pos)
    if lo == hi:
        return vals[lo]
    w = pos - lo
    return vals[lo] * (1 - w) + vals[hi] * w


def summarize(values: Sequence[float]) -> dict[str, float | int]:
    vals = [float(v) for v in values if not math.isnan(float(v))]
    if not vals:
        return {"n": 0, "mean": math.nan, "median": math.nan, "sd": math.nan, "p025": math.nan, "p975": math.nan}
    return {
        "n": len(vals),
        "mean": statistics.fmean(vals),
        "median": statistics.median(vals),
        "sd": statistics.stdev(vals) if len(vals) > 1 else 0.0,
        "p025": percentile(vals, 0.025),
        "p975": percentile(vals, 0.975),
    }


def summarize_rows(rows: Sequence[dict[str, Any]], group_fields: Sequence[str], metric_fields: Sequence[str]) -> list[dict[str, Any]]:
    groups: dict[tuple[Any, ...], dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
    examples: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in rows:
        key = tuple(row[f] for f in group_fields)
        examples[key] = {f: row[f] for f in group_fields}
        for metric in metric_fields:
            groups[key][metric].append(float(row[metric]))
    out = []
    for key, data in groups.items():
        for metric, values in data.items():
            out.append({**examples[key], "metric": metric, **summarize(values)})
    return out


def build_inext_input(
    out_path: Path,
    genera: Sequence[str],
    cells: Sequence[str],
    cell_masks: Sequence[int],
    band_indices: dict[str, list[int]],
    trait_masks: dict[str, int],
) -> None:
    rows: list[dict[str, Any]] = []
    # Band assemblages and pooled adjacent-pair assemblages. These are full retained
    # cell sets for coverage-standardized iNEXT diagnostics, not equal-cell draws.
    assemblages: list[tuple[str, str, list[int]]] = []
    for band in BAND_ORDER:
        assemblages.append(("band", band, band_indices[band]))
    for a, b in ADJACENT_PAIRS:
        assemblages.append(("adjacent_pair_pool", f"{a}__{b}", band_indices[a] + band_indices[b]))
    for scope, name, indices in assemblages:
        for trait in TRAIT_ORDER:
            tmask = trait_masks[trait]
            frequencies = [0] * len(genera)
            for ci in indices:
                masked = cell_masks[ci] & tmask
                while masked:
                    lsb = masked & -masked
                    gi = lsb.bit_length() - 1
                    frequencies[gi] += 1
                    masked ^= lsb
            for gi, freq in enumerate(frequencies):
                if freq > 0:
                    rows.append({
                        "scope": scope,
                        "assemblage": name,
                        "assemblage_label": BAND_LABEL.get(name, name.replace("__", " to ")),
                        "trait_class": trait,
                        "trait_label": TRAIT_LABEL[trait],
                        "n_sampling_units": len(indices),
                        "genus": genera[gi],
                        "incidence_frequency": freq,
                    })
    write_csv(out_path, rows, ["scope", "assemblage", "assemblage_label", "trait_class", "trait_label", "n_sampling_units", "genus", "incidence_frequency"])


def occurrence_audit(path: Path | None, genera: Sequence[str]) -> dict[str, Any]:
    if path is None:
        return {"available": False}
    fields, rows = read_delimited(path)
    genus_field = find_field(fields, ["analysis_genus", "genus", "trait_review_genus"], "occurrence genus", required=False)
    if genus_field is None:
        return {"available": True, "path": str(path), "rows": len(rows), "genus_field": None}
    genera_occ = {r.get(genus_field, "").strip().casefold() for r in rows if r.get(genus_field, "").strip()}
    matrix_set = {g.casefold() for g in genera}
    return {
        "available": True,
        "path": str(path),
        "rows": len(rows),
        "genus_field": genus_field,
        "unique_occurrence_genera": len(genera_occ),
        "matrix_genera_missing_from_occurrence": sorted(matrix_set - genera_occ),
        "occurrence_genera_not_in_matrix": sorted(genera_occ - matrix_set),
    }


def fmt(x: Any, digits: int = 3) -> str:
    try:
        value = float(x)
    except (TypeError, ValueError):
        return "NA"
    return "NA" if math.isnan(value) else f"{value:.{digits}f}"


def make_figures(output_dir: Path, richness_summary: list[dict[str, Any]], beta_summary: list[dict[str, Any]], contrasts: list[dict[str, Any]], sample_size: int, iterations: int) -> None:
    try:
        import matplotlib.pyplot as plt  # type: ignore
    except ImportError:
        (output_dir / "FIGURES_NOT_CREATED.txt").write_text("matplotlib is not installed. CSV analyses were completed.\n", encoding="utf-8")
        return
    figures = output_dir / "figures"
    figures.mkdir(parents=True, exist_ok=True)

    # Figure 1: aligned richness and adjacent-band Jaccard total.
    fig, axes = plt.subplots(2, 1, figsize=(12, 10), constrained_layout=True)
    x = list(range(len(BAND_ORDER)))
    for trait, marker in [("ballooning", "o"), ("non_ballooning", "s")]:
        rows = [r for r in richness_summary if r["trait_class"] == trait and r["metric"] == "richness"]
        lookup = {r["latitude_band"]: r for r in rows}
        y = [float(lookup[b]["median"]) for b in BAND_ORDER]
        lo = [y[i] - float(lookup[b]["p025"]) for i, b in enumerate(BAND_ORDER)]
        hi = [float(lookup[b]["p975"]) - y[i] for i, b in enumerate(BAND_ORDER)]
        axes[0].errorbar(x, y, yerr=[lo, hi], marker=marker, capsize=4, label=TRAIT_LABEL[trait])
    axes[0].set_xticks(x, [BAND_LABEL[b] for b in BAND_ORDER])
    axes[0].set_ylabel(f"Genus richness in {sample_size} occupied 25-km cells")
    axes[0].set_title("A. Equal-cell genus richness")
    axes[0].legend(frameon=False)

    pairs = ADJACENT_PAIRS
    xp = list(range(len(pairs)))
    for trait, marker in [("ballooning", "o"), ("non_ballooning", "s")]:
        rows = [r for r in beta_summary if r["trait_class"] == trait and r["metric"] == "jaccard_total" and r["adjacent"] is True]
        lookup = {(r["band_1"], r["band_2"]): r for r in rows}
        y = [float(lookup[p]["median"]) for p in pairs]
        lo = [y[i] - float(lookup[p]["p025"]) for i, p in enumerate(pairs)]
        hi = [float(lookup[p]["p975"]) - y[i] for i, p in enumerate(pairs)]
        axes[1].errorbar(xp, y, yerr=[lo, hi], marker=marker, capsize=4, label=TRAIT_LABEL[trait])
    axes[1].set_xticks(xp, [f"{BAND_LABEL[a]}\nto {BAND_LABEL[b]}" for a, b in pairs])
    axes[1].set_ylim(0, 1)
    axes[1].set_ylabel("Jaccard total dissimilarity")
    axes[1].set_title("B. Adjacent-band total compositional dissimilarity")
    axes[1].legend(frameon=False)
    fig.suptitle(f"Baja arachnid latitude-band QC ({iterations:,} paired resamples)")
    for ext in ("png", "pdf", "svg"):
        fig.savefig(figures / f"Figure_QC_01_richness_and_Jaccard.{ext}", dpi=300 if ext == "png" else None)
    plt.close(fig)

    # Figure 2: Baselga Jaccard replacement/nestedness decomposition.
    fig, axes = plt.subplots(2, 1, figsize=(13, 10), constrained_layout=True)
    width = 0.36
    for ax, trait in zip(axes, TRAIT_ORDER):
        turnover, nestedness = [], []
        for pair in pairs:
            t = next(r for r in beta_summary if r["trait_class"] == trait and r["metric"] == "jaccard_turnover" and (r["band_1"], r["band_2"]) == pair)
            n = next(r for r in beta_summary if r["trait_class"] == trait and r["metric"] == "jaccard_nestedness" and (r["band_1"], r["band_2"]) == pair)
            turnover.append(float(t["median"]))
            nestedness.append(float(n["median"]))
        ax.bar(xp, turnover, width=width, label="Baselga Jaccard turnover/replacement")
        ax.bar(xp, nestedness, width=width, bottom=turnover, label="Jaccard nestedness-resultant")
        ax.set_xticks(xp, [f"{BAND_LABEL[a]}\nto {BAND_LABEL[b]}" for a, b in pairs])
        ax.set_ylim(0, 1)
        ax.set_ylabel("Dissimilarity component")
        ax.set_title(TRAIT_LABEL[trait])
        ax.legend(frameon=False, loc="upper left")
    fig.suptitle("Baselga partition of adjacent-band Jaccard dissimilarity")
    for ext in ("png", "pdf", "svg"):
        fig.savefig(figures / f"Figure_QC_02_Baselga_Jaccard_partition.{ext}", dpi=300 if ext == "png" else None)
    plt.close(fig)

    # Figure 3: paired ballooning-minus-non-ballooning contrasts.
    fig, ax = plt.subplots(figsize=(13, 7), constrained_layout=True)
    metrics = ["jaccard_total", "jaccard_turnover", "jaccard_nestedness", "simpson_replacement", "sorensen_nestedness"]
    offsets = [(-0.28 + i * 0.14) for i in range(len(metrics))]
    for metric, offset in zip(metrics, offsets):
        rows = [r for r in contrasts if r["metric"] == metric and r["adjacent"] is True]
        lookup = {(r["band_1"], r["band_2"]): r for r in rows}
        y = [float(lookup[p]["median"]) for p in pairs]
        lo = [y[i] - float(lookup[p]["p025"]) for i, p in enumerate(pairs)]
        hi = [float(lookup[p]["p975"]) - y[i] for i, p in enumerate(pairs)]
        ax.errorbar([v + offset for v in xp], y, yerr=[lo, hi], marker="o", capsize=3, linestyle="none", label=metric.replace("_", " "))
    ax.axhline(0, linewidth=1, linestyle="--")
    ax.set_xticks(xp, [f"{BAND_LABEL[a]}\nto {BAND_LABEL[b]}" for a, b in pairs])
    ax.set_ylabel("Ballooning-capable minus non-ballooning")
    ax.set_title("Paired trait contrasts from identical cell draws")
    ax.legend(frameon=False, ncol=2)
    for ext in ("png", "pdf", "svg"):
        fig.savefig(figures / f"Figure_QC_03_trait_contrasts.{ext}", dpi=300 if ext == "png" else None)
    plt.close(fig)


def make_caption(output_dir: Path, beta_summary: list[dict[str, Any]], sample_size: int, iterations: int) -> None:
    lines = [
        "Figure caption draft (values inserted from the current run)",
        "",
        f"Genus-level diversity and compositional change across five latitude bands of the Baja California Peninsula. Each latitude band was standardized to {sample_size} occupied 25-km cells and resampled without replacement over {iterations:,} paired Monte Carlo iterations. Ballooning-capable and non-ballooning assemblages were reconstructed from the same cell draw in each iteration. Total Jaccard dissimilarity was partitioned following the Baselga replacement–nestedness framework into a turnover/replacement component and a nestedness-resultant component; the complementary Sørensen-family partition reports Simpson replacement and Sørensen nestedness-resultant. Values are medians with 2.5th–97.5th percentile resampling intervals.",
        "",
        "Adjacent-band medians:",
    ]
    for a, b in ADJACENT_PAIRS:
        lines.append(f"{BAND_LABEL[a]} to {BAND_LABEL[b]}:")
        for trait in TRAIT_ORDER:
            get = lambda metric: next(r for r in beta_summary if r["metric"] == metric and r["trait_class"] == trait and r["band_1"] == a and r["band_2"] == b)
            j = get("jaccard_total")
            jt = get("jaccard_turnover")
            jn = get("jaccard_nestedness")
            sim = get("simpson_replacement")
            sne = get("sorensen_nestedness")
            lines.append(
                f"  {TRAIT_LABEL[trait]}: Jaccard total {fmt(j['median'])} [{fmt(j['p025'])}, {fmt(j['p975'])}]; "
                f"Jaccard turnover {fmt(jt['median'])}; Jaccard nestedness {fmt(jn['median'])}; "
                f"Simpson replacement {fmt(sim['median'])}; Sørensen nestedness {fmt(sne['median'])}."
            )
    lines.extend([
        "",
        "Interpretive guardrail: Jaccard total dissimilarity is not equivalent to replacement. The Baselga and Sørensen partitions identify how much of total compositional difference is associated with balanced genus replacement versus richness-imbalance/nestedness-resultant differences. The iNEXT outputs are a separate coverage-standardized Hill-diversity diagnostic and should not be described as the Baselga partition.",
    ])
    (output_dir / "PUBLICATION_CAPTION_DRAFT.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Publication QC for Baja ballooning nestedness-versus-replacement analyses.")
    p.add_argument("project_root", nargs="?", default=None)
    p.add_argument("--iterations", type=int, default=DEFAULT_ITERATIONS)
    p.add_argument("--seed", type=int, default=DEFAULT_SEED)
    p.add_argument("--expected-equal-cells", type=int, default=DEFAULT_EXPECTED_EQUAL_CELLS)
    p.add_argument("--allow-cell-count-change", action="store_true")
    p.add_argument("--incidence")
    p.add_argument("--cell-lookup")
    p.add_argument("--traits")
    p.add_argument("--output-dir")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    if args.iterations < 100:
        raise ValueError("Use at least 100 iterations; publication run default is 5,000.")
    root = auto_project_root(args.project_root)
    inputs = discover_inputs(root, args)
    output_dir = Path(args.output_dir).expanduser().resolve() if args.output_dir else root / "04_analysis/11K_publication_nestedness_replacement_QC"
    output_dir.mkdir(parents=True, exist_ok=True)
    log_path = output_dir / "RUN_LOG.txt"

    class Tee:
        def __init__(self, *streams: Any): self.streams = streams
        def write(self, text: str) -> int:
            for s in self.streams:
                s.write(text); s.flush()
            return len(text)
        def flush(self) -> None:
            for s in self.streams: s.flush()

    with log_path.open("w", encoding="utf-8") as log:
        oldout, olderr = sys.stdout, sys.stderr
        sys.stdout = Tee(oldout, log)
        sys.stderr = Tee(olderr, log)
        try:
            print("BAJA BALLOONING PUBLICATION QC")
            print(f"Version: {SCRIPT_VERSION}")
            print(f"Started UTC: {utc_now()}")
            print(f"Project root: {root}")
            print(f"Iterations: {args.iterations:,}; seed: {args.seed}")
            print(f"Incidence matrix: {inputs.incidence_matrix}")
            print(f"Cell lookup: {inputs.cell_lookup}")
            print(f"Trait table: {inputs.trait_table}")

            incidence = read_incidence(inputs.incidence_matrix)
            genera, cells, cell_masks = incidence["genera"], incidence["cells"], incidence["cell_masks"]
            if "fesa" in {g.casefold() for g in genera}:
                raise RuntimeError("Known invalid genus 'Fesa' remains in the primary incidence matrix.")
            cell_bands, band_audit = read_cell_bands(inputs.cell_lookup, cells)
            traits = load_traits(inputs.trait_table, genera)
            trait_masks = {k: traits["masks"][k] for k in TRAIT_ORDER}

            band_indices = {band: [i for i, c in enumerate(cells) if cell_bands[c] == band] for band in BAND_ORDER}
            cell_counts = {band: len(v) for band, v in band_indices.items()}
            if any(v == 0 for v in cell_counts.values()):
                raise RuntimeError(f"At least one latitude band has no cells: {cell_counts}")
            sample_size = min(cell_counts.values())
            print(f"Genera in matrix: {len(genera)}")
            print(f"Occupied cells: {len(cells)}")
            print(f"Cells by band: {cell_counts}")
            print(f"Equal cells per band: {sample_size}")
            print(f"Trait evidence counts: {traits['evidence_counts']}")
            if sample_size != args.expected_equal_cells and not args.allow_cell_count_change:
                raise RuntimeError(
                    f"Equal-cell sample size changed from expected {args.expected_equal_cells} to {sample_size}. "
                    "This is a publication-QC stop. Inspect inputs or rerun with --allow-cell-count-change only after documenting the reason."
                )

            provenance = {
                "script_version": SCRIPT_VERSION,
                "started_utc": utc_now(),
                "project_root": str(root),
                "iterations": args.iterations,
                "seed": args.seed,
                "expected_equal_cells": args.expected_equal_cells,
                "actual_equal_cells": sample_size,
                "input_files": {
                    "incidence_matrix": {"path": str(inputs.incidence_matrix), "sha256": sha256(inputs.incidence_matrix), "bytes": inputs.incidence_matrix.stat().st_size},
                    "cell_lookup": {"path": str(inputs.cell_lookup), "sha256": sha256(inputs.cell_lookup), "bytes": inputs.cell_lookup.stat().st_size},
                    "trait_table": {"path": str(inputs.trait_table), "sha256": sha256(inputs.trait_table), "bytes": inputs.trait_table.stat().st_size, "sheet": traits["sheet"]},
                },
                "matrix_genera": len(genera),
                "matrix_cells": len(cells),
                "cell_counts_by_band": cell_counts,
                "trait_evidence_counts": traits["evidence_counts"],
                "trait_fields_used": traits["fields_used"],
                "occurrence_audit": occurrence_audit(inputs.occurrence_table, genera),
            }
            (output_dir / "00_INPUT_PROVENANCE.json").write_text(json.dumps(provenance, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
            write_csv(output_dir / "00_TRAIT_AUDIT.csv", traits["normalized_rows"], ["genus", "evidence_class", "analysis_class"])
            write_csv(output_dir / "00_CELL_BAND_AUDIT.csv", band_audit, ["cell_id", "raw_band", "normalized_band"])
            write_csv(output_dir / "00_BAND_CELL_COUNTS.csv", [{"latitude_band": b, "latitude_band_label": BAND_LABEL[b], "occupied_cells": cell_counts[b], "equal_cells_sampled": sample_size} for b in BAND_ORDER], ["latitude_band", "latitude_band_label", "occupied_cells", "equal_cells_sampled"])

            rng = random.Random(args.seed)
            richness_rows: list[dict[str, Any]] = []
            beta_rows: list[dict[str, Any]] = []
            contrast_rows: list[dict[str, Any]] = []
            pair_order = list(combinations(BAND_ORDER, 2))
            for iteration in range(1, args.iterations + 1):
                chosen = {band: rng.sample(band_indices[band], sample_size) for band in BAND_ORDER}
                union_by_band = {band: union_masks(cell_masks, chosen[band]) for band in BAND_ORDER}
                beta_by_pair_trait: dict[tuple[str, str, str], dict[str, float | int]] = {}
                for band in BAND_ORDER:
                    for trait in TRAIT_ORDER:
                        richness_rows.append({
                            "iteration": iteration,
                            "latitude_band": band,
                            "latitude_band_label": BAND_LABEL[band],
                            "trait_class": trait,
                            "trait_label": TRAIT_LABEL[trait],
                            "richness": (union_by_band[band] & trait_masks[trait]).bit_count(),
                        })
                for a, b in pair_order:
                    adjacent = (a, b) in ADJACENT_PAIRS
                    for trait in TRAIT_ORDER:
                        metrics = baselga_pair(union_by_band[a] & trait_masks[trait], union_by_band[b] & trait_masks[trait])
                        beta_by_pair_trait[(a, b, trait)] = metrics
                        beta_rows.append({
                            "iteration": iteration,
                            "band_1": a, "band_1_label": BAND_LABEL[a],
                            "band_2": b, "band_2_label": BAND_LABEL[b],
                            "adjacent": adjacent,
                            "trait_class": trait, "trait_label": TRAIT_LABEL[trait],
                            **metrics,
                        })
                    for metric in METRICS:
                        bv = float(beta_by_pair_trait[(a, b, "ballooning")][metric])
                        nv = float(beta_by_pair_trait[(a, b, "non_ballooning")][metric])
                        contrast_rows.append({
                            "iteration": iteration,
                            "band_1": a, "band_1_label": BAND_LABEL[a],
                            "band_2": b, "band_2_label": BAND_LABEL[b],
                            "adjacent": adjacent,
                            "metric": metric,
                            "ballooning_minus_non_ballooning": bv - nv if not math.isnan(bv) and not math.isnan(nv) else math.nan,
                        })
                if iteration % 500 == 0:
                    print(f"Completed {iteration:,}/{args.iterations:,} resampling iterations")

            richness_summary = summarize_rows(richness_rows, ["latitude_band", "latitude_band_label", "trait_class", "trait_label"], ["richness"])
            beta_summary = summarize_rows(beta_rows, ["band_1", "band_1_label", "band_2", "band_2_label", "adjacent", "trait_class", "trait_label"], ["shared_genera", "unique_to_band_1", "unique_to_band_2"] + METRICS)
            # Build contrast summaries explicitly so the ecological metric name is retained.
            contrast_summary: list[dict[str, Any]] = []
            grouped: dict[tuple[str, str, bool, str], list[float]] = defaultdict(list)
            for row in contrast_rows:
                grouped[(row["band_1"], row["band_2"], bool(row["adjacent"]), row["metric"])].append(float(row["ballooning_minus_non_ballooning"]))
            for (a, b, adjacent, metric), vals in grouped.items():
                contrast_summary.append({
                    "band_1": a, "band_1_label": BAND_LABEL[a], "band_2": b, "band_2_label": BAND_LABEL[b],
                    "adjacent": adjacent, "metric": metric, "contrast": "ballooning_minus_non_ballooning", **summarize(vals)
                })

            write_csv(output_dir / "01_EQUAL_CELL_RICHNESS_ITERATIONS.csv", richness_rows, ["iteration", "latitude_band", "latitude_band_label", "trait_class", "trait_label", "richness"])
            write_csv(output_dir / "02_BASELGA_PAIRWISE_ITERATIONS.csv", beta_rows, ["iteration", "band_1", "band_1_label", "band_2", "band_2_label", "adjacent", "trait_class", "trait_label", "shared_genera", "unique_to_band_1", "unique_to_band_2"] + METRICS)
            write_csv(output_dir / "03_TRAIT_CONTRAST_ITERATIONS.csv", contrast_rows, ["iteration", "band_1", "band_1_label", "band_2", "band_2_label", "adjacent", "metric", "ballooning_minus_non_ballooning"])
            write_csv(output_dir / "04_EQUAL_CELL_RICHNESS_SUMMARY.csv", richness_summary, ["latitude_band", "latitude_band_label", "trait_class", "trait_label", "metric", "n", "mean", "median", "sd", "p025", "p975"])
            write_csv(output_dir / "05_BASELGA_PAIRWISE_SUMMARY.csv", beta_summary, ["band_1", "band_1_label", "band_2", "band_2_label", "adjacent", "trait_class", "trait_label", "metric", "n", "mean", "median", "sd", "p025", "p975"])
            write_csv(output_dir / "06_TRAIT_CONTRAST_SUMMARY.csv", contrast_summary, ["band_1", "band_1_label", "band_2", "band_2_label", "adjacent", "metric", "contrast", "n", "mean", "median", "sd", "p025", "p975"])
            build_inext_input(output_dir / "07_iNEXT_INCIDENCE_FREQUENCY_INPUT.csv", genera, cells, cell_masks, band_indices, trait_masks)
            make_figures(output_dir, richness_summary, beta_summary, contrast_summary, sample_size, args.iterations)
            make_caption(output_dir, beta_summary, sample_size, args.iterations)

            # A compact table intended for manuscript-number checking.
            compact = []
            for row in beta_summary:
                if row["adjacent"] is True and row["metric"] in METRICS:
                    compact.append(row)
            write_csv(output_dir / "PUBLICATION_NUMBERS_ADJACENT_BANDS.csv", compact, ["band_1_label", "band_2_label", "trait_label", "metric", "median", "p025", "p975", "mean", "sd", "n"])

            report_lines = [
                "# Baja Ballooning publication QC report",
                "",
                f"Run completed from frozen inputs on {utc_now()}.",
                f"- Incidence genera: {len(genera)}",
                f"- Occupied 25-km cells: {len(cells)}",
                f"- Equal cells sampled per band: {sample_size}",
                f"- Paired Monte Carlo iterations: {args.iterations:,}",
                f"- Trait evidence counts: {traits['evidence_counts']}",
                "",
                "## Interpretation rule",
                "Jaccard total dissimilarity is partitioned into Baselga Jaccard turnover/replacement plus Jaccard nestedness-resultant. Simpson replacement is the turnover component of the Sørensen-family partition; it is related to, but not numerically identical to, Baselga Jaccard turnover.",
                "",
                "## Files to inspect first",
                "1. `00_INPUT_PROVENANCE.json`",
                "2. `PUBLICATION_NUMBERS_ADJACENT_BANDS.csv`",
                "3. `06_TRAIT_CONTRAST_SUMMARY.csv`",
                "4. `figures/Figure_QC_02_Baselga_Jaccard_partition.png`",
                "5. `PUBLICATION_CAPTION_DRAFT.txt`",
                "6. iNEXT outputs added by `02_run_iNEXT_hill.R` when R is available.",
            ]
            (output_dir / "QC_REPORT.md").write_text("\n".join(report_lines) + "\n", encoding="utf-8")
            print(f"Completed. Outputs: {output_dir}")
            return 0
        finally:
            sys.stdout, sys.stderr = oldout, olderr


if __name__ == "__main__":
    raise SystemExit(main())
