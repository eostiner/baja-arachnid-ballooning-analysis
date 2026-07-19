#!/usr/bin/env python3
"""Shared utilities for Phase 14 temporal recent-stress tracking analyses."""
from __future__ import annotations

import csv
import hashlib
import json
import math
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable, Iterator

SCRIPT_DATE = "2026-07-18"
EARTH_RADIUS_M = 6_371_008.8
DEFAULT_CENTER_LAT = 27.5
DEFAULT_CENTER_LON = -113.5
DEFAULT_CELL_SIZE_M = 25_000.0

GENUS_FIELDS = ("analysis_genus", "genus", "trait_review_genus", "accepted_genus", "final_genus", "Genus")
LAT_FIELDS = ("decimalLatitude", "latitude", "decimal_latitude")
LON_FIELDS = ("decimalLongitude", "longitude", "decimal_longitude")
ID_FIELDS = ("gbifID", "gbifId", "key", "occurrenceID", "occurrenceId")
DATE_FIELDS = ("eventDate", "event_date", "collectionDate", "collection_date")
YEAR_FIELDS = ("year", "eventYear", "collectionYear", "collection_year")
MONTH_FIELDS = ("month", "eventMonth", "collectionMonth", "collection_month")
DAY_FIELDS = ("day", "eventDay", "collectionDay", "collection_day")
EVENT_ID_FIELDS = ("eventID", "eventId", "samplingEventID", "sampling_event_id")
DATASET_FIELDS = (
    "datasetKey",
    "dataset_key",
    "datasetName",
    "datasetTitle",
    "publishingOrgKey",
    "institutionCode",
    "collectionCode",
)


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_genus(value: Any) -> str:
    text = clean(value)
    if not text:
        return ""
    return text[0].upper() + text[1:]


def first_field(fields: Iterable[str], candidates: Iterable[str], required: bool = False) -> str | None:
    exact = {field: field for field in fields}
    folded = {field.casefold(): field for field in fields}
    for candidate in candidates:
        if candidate in exact:
            return exact[candidate]
        if candidate.casefold() in folded:
            return folded[candidate.casefold()]
    if required:
        raise RuntimeError(
            "Required field not found. Tried: " + ", ".join(candidates) +
            ". Available: " + ", ".join(fields)
        )
    return None


def read_delimited(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    delimiter = "\t" if path.suffix.lower() in {".tsv", ".tab"} else ","
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as stream:
        reader = csv.DictReader(stream, delimiter=delimiter)
        return reader.fieldnames or [], list(reader)


def write_csv(path: Path, rows: Iterable[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    rows = list(rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        fieldnames = list(rows[0].keys()) if rows else []
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fieldnames, extrasaction="ignore", lineterminator="\n")
        if fieldnames:
            writer.writeheader()
            for row in rows:
                writer.writerow({key: row.get(key, "") for key in fieldnames})
    tmp.replace(path)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


@dataclass(frozen=True, order=True)
class CellKey:
    row: int
    col: int

    @property
    def cell_id(self) -> str:
        return f"BJA25K_C{self.col:+05d}_R{self.row:+05d}"


def forward_laea(lat_deg: float, lon_deg: float, center_lat: float, center_lon: float) -> tuple[float, float]:
    phi = math.radians(lat_deg)
    lam = math.radians(lon_deg)
    phi0 = math.radians(center_lat)
    lam0 = math.radians(center_lon)
    dlam = lam - lam0
    denominator = (
        1.0
        + math.sin(phi0) * math.sin(phi)
        + math.cos(phi0) * math.cos(phi) * math.cos(dlam)
    )
    if denominator <= 0:
        raise ValueError("Coordinate is antipodal to the projection center.")
    k = math.sqrt(2.0 / denominator)
    x = EARTH_RADIUS_M * k * math.cos(phi) * math.sin(dlam)
    y = EARTH_RADIUS_M * k * (
        math.cos(phi0) * math.sin(phi)
        - math.sin(phi0) * math.cos(phi) * math.cos(dlam)
    )
    return x, y


def cell_for_coordinate(
    lat: float,
    lon: float,
    cell_size_m: float = DEFAULT_CELL_SIZE_M,
    center_lat: float = DEFAULT_CENTER_LAT,
    center_lon: float = DEFAULT_CENTER_LON,
) -> CellKey:
    x, y = forward_laea(lat, lon, center_lat, center_lon)
    return CellKey(row=math.floor(y / cell_size_m), col=math.floor(x / cell_size_m))


def latitude_band(lat: float) -> str:
    if 23 <= lat < 24:
        return "23-24N"
    if 24 <= lat < 26:
        return "24-26N"
    if 26 <= lat < 28:
        return "26-28N"
    if 28 <= lat < 30:
        return "28-30N"
    if 30 <= lat <= 32.6:
        return "30-32N"
    return "outside_study_bands"


@dataclass(frozen=True)
class ParsedDate:
    year: int | None
    month: int | None
    day: int | None
    precision: str
    source: str
    status: str


_YEAR_RE = re.compile(r"(?<!\d)(1[5-9]\d{2}|20\d{2})(?!\d)")


def _int_or_none(value: Any) -> int | None:
    text = clean(value)
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def parse_occurrence_date(
    row: dict[str, str],
    date_field: str | None,
    year_field: str | None,
    month_field: str | None,
    day_field: str | None,
    current_year: int = 2026,
) -> ParsedDate:
    raw = clean(row.get(date_field)) if date_field else ""
    if raw:
        first = raw.split("/")[0].strip()
        # Excel serial dates occur in some preserved workbook exports.
        try:
            serial = float(first)
        except ValueError:
            serial = float("nan")
        if math.isfinite(serial) and 1 <= serial <= 100000:
            dt = datetime(1899, 12, 30) + timedelta(days=serial)
            if 1500 <= dt.year <= current_year:
                return ParsedDate(dt.year, dt.month, dt.day, "day_excel_serial", date_field or "", "valid")
        normalized = first.replace("Z", "+00:00")
        try:
            dt = datetime.fromisoformat(normalized)
            if 1500 <= dt.year <= current_year:
                return ParsedDate(dt.year, dt.month, dt.day, "day", date_field or "", "valid")
        except ValueError:
            pass
        # ISO-like YYYY-MM-DD / YYYY-MM / YYYY strings, with optional time suffix.
        match = re.match(r"^(\d{4})(?:-(\d{1,2}))?(?:-(\d{1,2}))?", first)
        if match:
            year = int(match.group(1))
            month = int(match.group(2)) if match.group(2) else None
            day = int(match.group(3)) if match.group(3) else None
            if 1500 <= year <= current_year and (month is None or 1 <= month <= 12):
                if day is not None:
                    try:
                        date(year, month or 1, day)
                    except ValueError:
                        day = None
                precision = "day" if day is not None else ("month" if month is not None else "year")
                return ParsedDate(year, month, day, precision, date_field or "", "valid")
        year_match = _YEAR_RE.search(first)
        if year_match:
            year = int(year_match.group(1))
            if 1500 <= year <= current_year:
                return ParsedDate(year, None, None, "year_from_text", date_field or "", "valid")

    year = _int_or_none(row.get(year_field)) if year_field else None
    month = _int_or_none(row.get(month_field)) if month_field else None
    day = _int_or_none(row.get(day_field)) if day_field else None
    if year is None:
        return ParsedDate(None, None, None, "missing", "", "missing")
    if not 1500 <= year <= current_year:
        return ParsedDate(None, None, None, "invalid", year_field or "", "invalid_year")
    if month is not None and not 1 <= month <= 12:
        month = None
    if day is not None:
        try:
            date(year, month or 1, day)
        except ValueError:
            day = None
    precision = "day" if day is not None else ("month" if month is not None else "year")
    return ParsedDate(year, month, day, precision, year_field or "", "valid")


def resolve_occurrence_file(project_root: Path, explicit: Path | None = None) -> Path:
    if explicit is not None:
        path = explicit.expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(f"Occurrence file not found: {path}")
        return path
    base = project_root / "02_data_clean" / "05_final_qc_flags"
    candidates = [
        base / "05_biodiversity_final_records.tsv",
        base / "05_ballooning_final_records.tsv",
        base / "05_environmental_temporal_sensitivity_2001_2020.tsv",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise FileNotFoundError(
        "No retained final-QC occurrence table found. Tried:\n" +
        "\n".join(str(candidate) for candidate in candidates)
    )


def resolve_trait_file(project_root: Path, explicit: Path | None = None) -> Path:
    if explicit is not None:
        path = explicit.expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(f"Trait file not found: {path}")
        return path
    candidates = [
        project_root / "ANALYSIS_READY_INPUTS" / "03_trait_tables" / "07_reviewed_genus_trait_lookup_normalized.csv",
        project_root / "ANALYSIS_READY_INPUTS" / "03_trait_tables" / "07_reviewed_genus_trait_lookup_final.csv",
        project_root / "02_data_clean" / "07_final_trait_merge" / "07_reviewed_genus_trait_lookup_final.csv",
        project_root / "USE_GOOD_BalloonID_Baja_Arachnid_GenusSpecies_Long_D1_D4_AUTHORITATIVE.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise FileNotFoundError(
        "No authoritative trait table found. Tried:\n" +
        "\n".join(str(candidate) for candidate in candidates)
    )


def _trait_group_from_values(row: dict[str, Any]) -> tuple[str, str]:
    analysis_candidates = (
        "analysis_class", "primary_C3_group", "primary_c3_group", "trait_group",
        "ballooning_group", "Ballooning_Behavior",
    )
    tier_candidates = (
        "evidence_class", "exclusive_tier", "final_tier_for_current_build", "tier",
    )
    analysis_text = ""
    tier_text = ""
    folded = {str(key).casefold(): key for key in row}
    for candidate in analysis_candidates:
        key = folded.get(candidate.casefold())
        if key is not None and clean(row.get(key)):
            analysis_text = clean(row.get(key))
            break
    for candidate in tier_candidates:
        key = folded.get(candidate.casefold())
        if key is not None and clean(row.get(key)):
            tier_text = clean(row.get(key)).upper()
            break
    normalized = re.sub(r"[^a-z0-9]+", "", analysis_text.casefold())
    if normalized in {"c3", "ballooning", "ballooningc3", "primaryc3", "yes"}:
        return "C3", tier_text or "C3"
    if normalized in {"n0", "nonballooning", "nonballooningn0", "fixedn0", "no"}:
        return "N0", tier_text or "N0"
    if "excluded" in normalized or normalized == "d4":
        return "EXCLUDED_D4", tier_text or "D4"
    if tier_text in {"D1", "D2", "D3"}:
        return "C3", tier_text
    if tier_text == "D4":
        return "EXCLUDED_D4", tier_text
    if tier_text == "N0":
        return "N0", tier_text
    return "UNRESOLVED", tier_text


def load_trait_lookup(path: Path) -> dict[str, dict[str, str]]:
    rows: list[dict[str, Any]]
    fields: list[str]
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is required to read the authoritative Excel trait workbook. "
                "Run `python -m pip install openpyxl` or provide the normalized CSV."
            ) from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        preferred = "Genus_Trait_Master_267"
        sheet = workbook[preferred] if preferred in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        iterator = sheet.iter_rows(values_only=True)
        header = [clean(value) for value in next(iterator)]
        fields = header
        rows = [dict(zip(header, values)) for values in iterator]
    else:
        fields, rows = read_delimited(path)
    genus_field = first_field(fields, GENUS_FIELDS, required=True)
    lookup: dict[str, dict[str, str]] = {}
    duplicates: set[str] = set()
    for row in rows:
        genus = normalize_genus(row.get(genus_field))
        if not genus:
            continue
        group, evidence = _trait_group_from_values(row)
        if genus in lookup:
            duplicates.add(genus)
        lookup[genus] = {
            "trait_group": group,
            "evidence_class": evidence,
        }
    if duplicates:
        raise RuntimeError("Duplicate genera in trait table: " + ", ".join(sorted(duplicates)))
    return lookup


def load_periods(path: Path) -> list[dict[str, Any]]:
    fields, rows = read_delimited(path)
    required = {"period_id", "start_year", "end_year"}
    if not required.issubset(fields):
        raise RuntimeError(f"Temporal-window config missing columns: {sorted(required - set(fields))}")
    periods = []
    for row in rows:
        period_id = clean(row.get("period_id"))
        start = int(clean(row.get("start_year")))
        end = int(clean(row.get("end_year")))
        if not period_id or end < start:
            raise RuntimeError(f"Invalid period row: {row}")
        periods.append({"period_id": period_id, "start_year": start, "end_year": end})
    periods.sort(key=lambda item: (item["start_year"], item["end_year"], item["period_id"]))
    for previous, current in zip(periods, periods[1:]):
        if current["start_year"] <= previous["end_year"]:
            raise RuntimeError("Temporal-window config contains overlapping periods.")
    return periods


def period_for_year(year: int, periods: list[dict[str, Any]]) -> str | None:
    for period in periods:
        if period["start_year"] <= year <= period["end_year"]:
            return str(period["period_id"])
    return None


def default_analysis_output_root(project_root: Path) -> Path:
    preferred = project_root / "04_analysis_USE _THIS"
    if preferred.exists():
        return preferred / "14_recent_environmental_stress"
    return project_root / "04_analysis" / "14_recent_environmental_stress"


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, default=str) + "\n", encoding="utf-8")


def quantile(values: list[float], q: float) -> float:
    if not values:
        return float("nan")
    ordered = sorted(values)
    if len(ordered) == 1:
        return ordered[0]
    position = (len(ordered) - 1) * q
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return ordered[lower]
    weight = position - lower
    return ordered[lower] * (1 - weight) + ordered[upper] * weight


def iter_adjacent_period_pairs(periods: list[dict[str, Any]]) -> Iterator[tuple[str, str]]:
    for left, right in zip(periods, periods[1:]):
        yield str(left["period_id"]), str(right["period_id"])

_CELL_ID_RE = re.compile(r"^BJA25K_C(?P<col>[+-]\d+)_R(?P<row>[+-]\d+)$")


def parse_cell_id(cell_id: str) -> CellKey:
    """Parse a retained BJA25K grid-cell identifier into integer row/column indices."""
    match = _CELL_ID_RE.match(clean(cell_id))
    if not match:
        raise ValueError(f"Invalid BJA25K grid cell id: {cell_id!r}")
    return CellKey(row=int(match.group("row")), col=int(match.group("col")))


def inverse_laea(
    x_m: float,
    y_m: float,
    center_lat: float = DEFAULT_CENTER_LAT,
    center_lon: float = DEFAULT_CENTER_LON,
) -> tuple[float, float]:
    """Inverse spherical Lambert azimuthal equal-area projection used by the retained grid."""
    phi0 = math.radians(center_lat)
    lam0 = math.radians(center_lon)
    rho = math.hypot(x_m, y_m)
    if rho == 0:
        return center_lat, center_lon
    ratio = min(1.0, rho / (2.0 * EARTH_RADIUS_M))
    c = 2.0 * math.asin(ratio)
    sinc = math.sin(c)
    cosc = math.cos(c)
    phi = math.asin(
        cosc * math.sin(phi0)
        + (y_m * sinc * math.cos(phi0) / rho)
    )
    lam = lam0 + math.atan2(
        x_m * sinc,
        rho * math.cos(phi0) * cosc - y_m * math.sin(phi0) * sinc,
    )
    return math.degrees(phi), math.degrees(lam)


def cell_polygon_lonlat(
    cell_id: str,
    cell_size_m: float = DEFAULT_CELL_SIZE_M,
    center_lat: float = DEFAULT_CENTER_LAT,
    center_lon: float = DEFAULT_CENTER_LON,
    segments_per_edge: int = 8,
) -> list[list[float]]:
    """Return a densified lon/lat ring representing the retained 25-km LAEA cell."""
    if segments_per_edge < 1:
        raise ValueError("segments_per_edge must be at least 1")
    key = parse_cell_id(cell_id)
    x0 = key.col * cell_size_m
    x1 = (key.col + 1) * cell_size_m
    y0 = key.row * cell_size_m
    y1 = (key.row + 1) * cell_size_m
    corners = [(x0, y0), (x1, y0), (x1, y1), (x0, y1)]
    boundary_xy: list[tuple[float, float]] = []
    for (xa, ya), (xb, yb) in zip(corners, corners[1:] + corners[:1]):
        for step in range(segments_per_edge):
            fraction = step / segments_per_edge
            boundary_xy.append((xa + fraction * (xb - xa), ya + fraction * (yb - ya)))
    boundary_xy.append(boundary_xy[0])
    ring: list[list[float]] = []
    for x_m, y_m in boundary_xy:
        lat, lon = inverse_laea(x_m, y_m, center_lat=center_lat, center_lon=center_lon)
        ring.append([lon, lat])
    return ring
