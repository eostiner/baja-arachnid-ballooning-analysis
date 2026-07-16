#!/usr/bin/env python3
"""Normalize the authoritative genus trait table for the publication pipeline.

The audited workbook/export uses two complementary fields:

- ``exclusive_tier``: D1, D2, D3, or D4 for genera with ballooning evidence.
- ``primary_C3_group``: C3/Ballooning, N0/Non-ballooning, or D4 excluded.

Non-ballooning genera may legitimately have a blank ``exclusive_tier``.  This
script combines both fields, validates their consistency, and writes a single
canonical downstream table with explicit ``evidence_class`` and
``analysis_class`` columns.  The source table is never modified.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

TIER_TOKEN = re.compile(r"(?<![A-Z0-9])(D1|D2|D3|D4|N0|C3)(?![A-Z0-9])", re.I)

GENUS_FIELDS = ("genus", "analysis_genus", "accepted_genus", "final_genus")
TIER_FIELDS = (
    "exclusive_tier",
    "final_tier_for_current_build",
    "evidence_class",
    "final_evidence_class",
    "final_evidence_category",
    "evidence_category",
    "ballooning_evidence_tier",
    "ballooning_evidence_category",
    "evidence_tier",
    "evidence_level",
    "trait_evidence_level",
    "d_level",
    "dlevel",
)
GROUP_FIELDS = (
    "primary_C3_group",
    "analysis_class",
    "c3_analysis_class",
    "primary_class",
    "trait_class",
    "final_designation",
    "designation",
)
CONFIDENCE_FIELDS = (
    "final_confidence",
    "trait_final_confidence",
    "trait_confidence",
    "trait_ballooning_confidence",
)
ORDER_FIELDS = ("order", "taxon_order", "order_final")
FAMILY_FIELDS = ("family", "taxon_family", "family_final")
AUTHORITATIVE_WORKBOOK_NAMES = (
    "USE_GOOD_BalloonID_Baja_Arachnid_GenusSpecies_Long_D1_D4_AUTHORITATIVE.xlsx",
    "USE_GOOD_BalloonID_Baja_Arachnid_Species_Long_REAUDITED.xlsx",
)


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalized(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean(value).casefold())


def field_lookup(fields: Iterable[str]) -> dict[str, str]:
    return {field.casefold(): field for field in fields}


def first_field(fields: list[str], candidates: Iterable[str], required: bool = False) -> str | None:
    lower = field_lookup(fields)
    for candidate in candidates:
        if candidate.casefold() in lower:
            return lower[candidate.casefold()]
    if required:
        raise ValueError(
            "Could not find a required field. Tried: " + ", ".join(candidates)
            + ". Available: " + ", ".join(fields)
        )
    return None


def parse_tier(value: Any) -> str | None:
    text = clean(value)
    if not text:
        return None
    tokens = {token.upper() for token in TIER_TOKEN.findall(text.upper())}
    if len(tokens) == 1:
        return next(iter(tokens))
    n = normalized(text)
    if n in {
        "nonballooning", "fixednonballooning", "referencenonballooning",
        "nonballooningreference", "noballooning", "n0reference",
    }:
        return "N0"
    if n in {"c3", "primaryc3", "d1d2d3", "d1tod3", "ballooningc3"}:
        return "C3"
    if n in {"d4excluded", "excludedd4", "d4broadinference"}:
        return "D4"
    return None


def parse_group(value: Any) -> str | None:
    text = clean(value)
    if not text:
        return None
    upper = text.upper()
    tokens = {token.upper() for token in TIER_TOKEN.findall(upper)}
    n = normalized(text)

    # Exclusion must be evaluated before generic "ballooning" text.
    if "D4" in tokens or ("d4" in n and ("exclude" in n or "broad" in n)):
        return "D4_excluded"
    if "N0" in tokens or "nonballoon" in n or n in {
        "fixednonballooning", "referencenonballooning", "noballooning",
        "nonballooningreference", "reference", "negative",
    }:
        return "N0"
    if tokens & {"D1", "D2", "D3", "C3"}:
        return "C3"
    if n in {
        "c3", "primaryc3", "ballooning", "ballooningc3", "c3ballooning",
        "positive", "included", "primaryballooning",
    }:
        return "C3"
    if "exclude" in n:
        return "D4_excluded"
    return None


def group_from_tier(tier: str | None) -> str | None:
    if tier in {"D1", "D2", "D3", "C3"}:
        return "C3"
    if tier == "N0":
        return "N0"
    if tier == "D4":
        return "D4_excluded"
    return None


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def find_source(root: Path) -> Path:
    workbook_candidates = [
        *(root / name for name in AUTHORITATIVE_WORKBOOK_NAMES),
        *(root / "ANALYSIS_READY_INPUTS/03_trait_tables" / name for name in AUTHORITATIVE_WORKBOOK_NAMES),
        *(root / "02_data_clean/07_final_trait_merge" / name for name in AUTHORITATIVE_WORKBOOK_NAMES),
    ]
    csv_candidates = [
        root / "ANALYSIS_READY_INPUTS/03_trait_tables/07_reviewed_genus_trait_lookup_final.csv",
        root / "02_data_clean/07_final_trait_merge/07_reviewed_genus_trait_lookup_final.csv",
    ]
    candidates = workbook_candidates + csv_candidates
    for path in candidates:
        if path.is_file():
            return path
    raise FileNotFoundError("Could not find the authoritative trait table. Tried:\n" + "\n".join(map(str, candidates)))


def read_source_table(source: Path) -> tuple[list[dict[str, Any]], list[str], str | None]:
    if source.suffix.casefold() == ".csv":
        with source.open(encoding="utf-8-sig", newline="") as stream:
            reader = csv.DictReader(stream)
            return list(reader), list(reader.fieldnames or []), None

    if source.suffix.casefold() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Reading the authoritative workbook requires openpyxl>=3.1") from exc
        workbook = load_workbook(source, read_only=True, data_only=True)
        preferred = "Genus_Trait_Master_267"
        sheet_names = [preferred] + [name for name in workbook.sheetnames if name != preferred]
        for sheet_name in sheet_names:
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            values = sheet.iter_rows(values_only=True)
            try:
                header_values = next(values)
            except StopIteration:
                continue
            fields = [clean(value) for value in header_values]
            if not any(fields):
                continue
            if first_field(fields, GENUS_FIELDS) is None:
                continue
            if first_field(fields, TIER_FIELDS) is None and first_field(fields, GROUP_FIELDS) is None:
                continue
            rows = []
            for values_row in values:
                row = {field: value for field, value in zip(fields, values_row) if field}
                if any(clean(value) for value in row.values()):
                    rows.append(row)
            if rows:
                return rows, fields, sheet_name
        raise ValueError(
            f"No workbook sheet contains a genus field and an authoritative tier/group field: {source}"
        )

    raise ValueError(f"Unsupported trait-table format: {source.suffix}")


def matrix_genera(root: Path) -> set[str] | None:
    candidates = [
        root / "ANALYSIS_READY_INPUTS/02_incidence_matrices_25km/10_biodiversity_final_genus_by_grid25km_incidence.csv",
        root / "02_data_clean/08_grid25km_incidence/10_biodiversity_final_genus_by_grid25km_incidence.csv",
    ]
    path = next((p for p in candidates if p.is_file()), None)
    if path is None:
        return None
    with path.open(encoding="utf-8-sig", newline="") as stream:
        reader = csv.DictReader(stream)
        fields = reader.fieldnames or []
        genus_field = first_field(fields, GENUS_FIELDS, required=True)
        return {clean(row.get(genus_field, "")).casefold() for row in reader if clean(row.get(genus_field, ""))}


@dataclass
class NormalizationResult:
    source: Path
    output: Path
    tier_field: str | None
    group_field: str | None
    genus_count: int
    counts: Counter[str]
    evidence_counts: Counter[str]


def normalize_project_traits(root: Path, *, write: bool = True) -> NormalizationResult:
    root = root.expanduser().resolve()
    source = find_source(root)
    rows, fields, source_sheet = read_source_table(source)
    if not rows:
        raise ValueError(f"Trait table is empty: {source}")

    genus_field = first_field(fields, GENUS_FIELDS, required=True)
    tier_field = first_field(fields, TIER_FIELDS)
    group_field = first_field(fields, GROUP_FIELDS)
    confidence_field = first_field(fields, CONFIDENCE_FIELDS)
    order_field = first_field(fields, ORDER_FIELDS)
    family_field = first_field(fields, FAMILY_FIELDS)

    if tier_field is None and group_field is None:
        raise ValueError(
            "Trait table has neither an evidence-tier field nor a primary C3/N0 group field. "
            "Expected columns such as exclusive_tier and primary_C3_group."
        )

    output_rows: list[dict[str, str]] = []
    seen: dict[str, int] = {}
    conflicts: list[str] = []
    unresolved: list[str] = []

    for line_number, row in enumerate(rows, start=2):
        genus = clean(row.get(genus_field, ""))
        if not genus:
            continue
        key = genus.casefold()
        if key in seen:
            raise ValueError(f"Duplicate genus {genus!r} on lines {seen[key]} and {line_number}.")
        seen[key] = line_number

        tier = parse_tier(row.get(tier_field, "")) if tier_field else None
        group = parse_group(row.get(group_field, "")) if group_field else None
        derived = group_from_tier(tier)

        if group is not None and derived is not None and group != derived:
            conflicts.append(
                f"{genus}: {group_field}={row.get(group_field, '')!r} -> {group}; "
                f"{tier_field}={row.get(tier_field, '')!r} -> {tier}"
            )
            continue

        analysis_class = group or derived
        if analysis_class is None:
            unresolved.append(
                f"{genus}: {group_field or '<no group field>'}={row.get(group_field, '') if group_field else ''!r}; "
                f"{tier_field or '<no tier field>'}={row.get(tier_field, '') if tier_field else ''!r}"
            )
            continue

        if tier is None:
            tier = {"C3": "C3", "N0": "N0", "D4_excluded": "D4"}[analysis_class]

        primary_group = {
            "C3": "Ballooning (C3)",
            "N0": "Non-ballooning (N0)",
            "D4_excluded": "Excluded (D4)",
        }[analysis_class]

        output_rows.append({
            "genus": genus,
            "order": clean(row.get(order_field, "")) if order_field else "",
            "family": clean(row.get(family_field, "")) if family_field else "",
            "exclusive_tier": tier,
            "primary_C3_group": primary_group,
            "evidence_class": tier,
            "analysis_class": analysis_class,
            "ballooning_binary_primary": "1" if analysis_class == "C3" else "0" if analysis_class == "N0" else "",
            "final_confidence": (clean(row.get(confidence_field, "")) if confidence_field else "UNSPECIFIED") or "UNSPECIFIED",
            "source_tier_field": tier_field or "",
            "source_group_field": group_field or "",
        })

    if conflicts:
        raise ValueError("Conflicting primary-group and tier designations:\n" + "\n".join(conflicts[:30]))
    if unresolved:
        raise ValueError("Unresolved trait designations:\n" + "\n".join(unresolved[:30]))
    if not output_rows:
        raise ValueError("No trait rows could be normalized.")

    matrix = matrix_genera(root)
    if matrix is not None:
        table = {row["genus"].casefold() for row in output_rows}
        missing = sorted(matrix - table)
        if missing:
            raise ValueError(
                f"{len(missing)} incidence-matrix genera are absent from the normalized trait table: "
                + ", ".join(missing[:30])
            )

    counts = Counter(row["analysis_class"] for row in output_rows)
    evidence_counts = Counter(row["evidence_class"] for row in output_rows)
    if counts["C3"] == 0 or counts["N0"] == 0:
        raise ValueError(
            f"Primary classes are invalid: C3={counts['C3']}, N0={counts['N0']}, "
            f"D4_excluded={counts['D4_excluded']}."
        )

    output = root / "ANALYSIS_READY_INPUTS/03_trait_tables/07_reviewed_genus_trait_lookup_normalized.csv"
    if write:
        output.parent.mkdir(parents=True, exist_ok=True)
        with output.open("w", encoding="utf-8", newline="") as stream:
            writer = csv.DictWriter(stream, fieldnames=list(output_rows[0]))
            writer.writeheader()
            writer.writerows(output_rows)
        provenance = {
            "source": str(source),
            "source_sha256": sha256(source),
            "output": str(output),
            "output_sha256": sha256(output),
            "genus_count": len(output_rows),
            "tier_field": tier_field,
            "group_field": group_field,
            "source_sheet": source_sheet,
            "analysis_class_counts": dict(sorted(counts.items())),
            "evidence_class_counts": dict(sorted(evidence_counts.items())),
            "primary_definition": "C3 = D1 + D2 + D3 versus fixed N0; D4 excluded",
        }
        with output.with_suffix(".provenance.json").open("w", encoding="utf-8") as stream:
            json.dump(provenance, stream, indent=2)
            stream.write("\n")

    return NormalizationResult(
        source=source,
        output=output,
        tier_field=tier_field,
        group_field=group_field,
        genus_count=len(output_rows),
        counts=counts,
        evidence_counts=evidence_counts,
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("project_root", type=Path)
    parser.add_argument("--check-only", action="store_true", help="Validate without writing the normalized table.")
    args = parser.parse_args()
    result = normalize_project_traits(args.project_root, write=not args.check_only)
    print(f"TRAIT SOURCE: {result.source}")
    print(f"TIER FIELD: {result.tier_field or '<derived from primary group>'}")
    print(f"PRIMARY GROUP FIELD: {result.group_field or '<derived from tier>'}")
    print(
        "PRIMARY CLASSES: "
        f"C3={result.counts['C3']}, N0={result.counts['N0']}, "
        f"D4_excluded={result.counts['D4_excluded']}"
    )
    print("EVIDENCE COUNTS: " + ", ".join(f"{k}={v}" for k, v in sorted(result.evidence_counts.items())))
    if not args.check_only:
        print(f"NORMALIZED TRAIT TABLE: {result.output}")
    print("TRAIT NORMALIZATION PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
